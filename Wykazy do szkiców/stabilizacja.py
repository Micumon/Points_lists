import openpyxl

path = "C:\\Users\\admin\\Desktop\\stabilizacja\\"

stabilization = {}
with open(path+"rep0.txt") as file:
    for line in file:
        line = line.strip().split("\t")
        stabilization.update({line[0].strip(): line[1].strip()})
print(stabilization)

wb_obj = openpyxl.load_workbook(path+"punkty graniczne Oryszew Osada stab.xlsx")
sheet = wb_obj.active

rows = sheet.max_row
for row in range(2, rows+1):
    if sheet.cell(row=row, column=8).value is None:
        sheet.cell(row=row, column=8).value = stabilization[str(sheet.cell(row=row, column=1).value).replace("obl", "")
        .replace("pom", "")]
    print(sheet.cell(row=row, column=1).value, "  ", sheet.cell(row=row, column=8).value)

wb_obj.save(path+"punkty graniczne Oryszew Osada stab.xlsx")
